#!/usr/bin/env python3
"""
SLO Analysis Script

A specialized script for analyzing SLOs with comprehensive user and service information.
Leverages code from account_analyzer.py and users_basic_v1.2.py to:
- Authenticate and select context (supports custom instances like us1, cg1)
- Collect SLOs and services using sloctl
- Fetch user information to resolve created_by user IDs and service responsible users
- Export to Excel with:
  - SLO information including user names, emails, and responsible users
  - Service information with SLO counts and responsible users
  - Hyperlinks to SLOs and services in the Nobl9 UI
  - Hidden columns for technical details (queries, targets, etc.)
"""

import json
import sys
import subprocess
import argparse
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path

try:
    import toml
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Alignment
    import colorama
    colorama.init()
except ImportError as e:
    print(f"Missing required dependencies: {e}")
    print("Install with: pip install toml pandas openpyxl colorama")
    sys.exit(1)

# Import from account_analyzer (same directory)
try:
    from account_analyzer import (
        enhanced_choose_context,
        get_credentials_for_context,
        authenticate,
        SLOInfo
    )
except ImportError as e:
    print(f"Error importing from account_analyzer: {e}")
    sys.exit(1)

# Import user fetching functions from users_basic_v1.2
# Use importlib to load module with dots in filename
import importlib.util
user_scripts_path = Path(__file__).parent.parent / "user_scripts" / "users_basic_v1.2.py"
if not user_scripts_path.exists():
    print(f"ERROR: users_basic_v1.2.py not found at {user_scripts_path}")
    sys.exit(1)

spec = importlib.util.spec_from_file_location("users_basic_v1_2", user_scripts_path)
if spec and spec.loader:
    users_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(users_module)
    fetch_users = users_module.fetch_users
else:
    print("ERROR: Could not load users_basic_v1.2.py")
    sys.exit(1)


def print_colored(text: str, color: str = "") -> None:
    """Print colored text to console.
    
    Args:
        text: Text to print
        color: Colorama color code (e.g., colorama.Fore.GREEN)
    """
    print(f"{color}{text}{colorama.Fore.RESET}")


def print_header(text: str) -> None:
    """Print a formatted header section.
    
    Args:
        text: Header text to display
    """
    print_colored(f"\n{'='*60}", colorama.Fore.CYAN)
    print_colored(text, colorama.Fore.CYAN)
    print_colored('='*60, colorama.Fore.CYAN)


class SLOAnalyzer:
    """SLO Analyzer with user information resolution."""
    
    def __init__(self, context_name: Optional[str] = None):
        """Initialize the analyzer."""
        self.context_name = context_name
        self.access_token = None
        self.organization_id = None
        self.base_url = "https://app.nobl9.com"
        self.is_custom_instance = False
        self.slos: List[SLOInfo] = []
        self.composite_slos: List[SLOInfo] = []
        self.services: List[Dict] = []
        self.user_lookup: Dict[str, Dict[str, str]] = {}  # user_id -> {name, email}
        self.service_responsible_users: Dict[Tuple[str, str], List[str]] = {}  # (project, service_name) -> [user_ids]
        self.service_slo_counts: Dict[Tuple[str, str], int] = {}  # (project, service_name) -> SLO count
        
    def setup_authentication(self):
        """Setup authentication and context."""
        print_header("AUTHENTICATION")
        
        if self.context_name:
            print_colored(f"Using specified context: {self.context_name}", colorama.Fore.GREEN)
            credentials = get_credentials_for_context(self.context_name)
            if not credentials:
                print_colored(f"Context '{self.context_name}' not found", colorama.Fore.RED)
                sys.exit(1)
        else:
            self.context_name, credentials = enhanced_choose_context()
            print_colored(f"Selected context: {self.context_name}", colorama.Fore.GREEN)
        
        token, org_id, is_custom_instance, base_url = authenticate(credentials)
        print_colored("✓ Authentication successful", colorama.Fore.GREEN)
        
        self.access_token = token
        self.organization_id = org_id
        self.is_custom_instance = is_custom_instance
        
        if is_custom_instance and base_url:
            self.base_url = base_url
        
        # Switch sloctl context
        print_colored(f"\nSwitching to context: {self.context_name}", colorama.Fore.CYAN)
        try:
            subprocess.run(
                ["sloctl", "config", "use-context", self.context_name],
                capture_output=True, text=True, check=True
            )
            print_colored("✓ Context switched successfully", colorama.Fore.GREEN)
        except subprocess.CalledProcessError as e:
            print_colored(f"ERROR: Failed to switch context: {e}", colorama.Fore.RED)
            sys.exit(1)
        
        print_colored(f"Organization: {self.organization_id}", colorama.Fore.GREEN)
    
    def _run_sloctl_command(self, command: List[str]) -> List[Dict]:
        """Run sloctl command and return JSON output."""
        try:
            result = subprocess.run(
                ["sloctl"] + command + ["-o", "json"],
                capture_output=True, text=True, check=True
            )
            
            if not result.stdout.strip():
                return []
            
            data = json.loads(result.stdout)
            return data if isinstance(data, list) else [data]
        except (subprocess.CalledProcessError, json.JSONDecodeError) as e:
            print_colored(f"Error running sloctl command: {e}", colorama.Fore.YELLOW)
            return []
    
    def collect_slos(self):
        """Collect all SLOs using sloctl."""
        print_header("COLLECTING SLOS")
        print("Collecting SLOs...")
        
        data = self._run_sloctl_command(["get", "slos", "-A"])
        
        if not data or not isinstance(data, list):
            print_colored("No SLO data found", colorama.Fore.YELLOW)
            return
        
        # Parse SLOs using similar logic to account_analyzer
        for item in data:
            metadata = item.get("metadata", {})
            spec = item.get("spec", {})
            objectives = spec.get("objectives", [])
            
            # Check if composite SLO
            is_composite = False
            for objective in objectives:
                if "composite" in objective and "components" in objective.get("composite", {}):
                    is_composite = True
                    break
            
            # Get target
            slo_target = 0.0
            if objectives and not is_composite:
                for objective in objectives:
                    if "target" in objective:
                        slo_target = objective.get("target", 0.0)
                        break
            else:
                slo_target = spec.get("target", 0.0)
            
            # Extract queries (skip for composite SLOs)
            query, numerator_query, denominator_query = "", "", ""
            if not is_composite:
                query, numerator_query, denominator_query = self._extract_slo_queries(spec)
            
            slo = SLOInfo(
                name=metadata.get("name", ""),
                project=metadata.get("project", ""),
                service=spec.get("service", ""),
                description=spec.get("description", ""),
                target=slo_target,
                time_window=str(spec.get("timeWindows", [])),
                alert_policies=spec.get("alertPolicies", []),
                health_status=item.get("status", {}).get("health", "unknown"),
                created_at=spec.get("createdAt", ""),
                updated_at=item.get("status", {}).get("updatedAt", ""),
                display_name=metadata.get("displayName", ""),
                query=query,
                numerator_query=numerator_query,
                denominator_query=denominator_query,
                created_by=spec.get("createdBy", "")
            )
            
            # Store SLO type as an attribute for export
            slo.slo_type = "Composite" if is_composite else "Regular"
            
            # Add to appropriate list
            if is_composite:
                self.composite_slos.append(slo)
            else:
                self.slos.append(slo)
        
        # Count SLOs per service
        self._count_slos_per_service()
        
        # Print summary
        print_colored(f"✓ Collected {len(self.slos)} regular SLOs", colorama.Fore.GREEN)
        print_colored(f"✓ Collected {len(self.composite_slos)} composite SLOs", colorama.Fore.GREEN)
        print_colored("-" * 60, colorama.Fore.CYAN)
        total_slos = len(self.slos) + len(self.composite_slos)
        print_colored(f"Total: {total_slos} SLOs", colorama.Fore.GREEN)
    
    def _count_slos_per_service(self) -> None:
        """Count the number of SLOs for each service (including composites)."""
        self.service_slo_counts.clear()
        # Count both regular and composite SLOs
        all_slos = self.slos + self.composite_slos
        for slo in all_slos:
            if slo.project and slo.service:
                key = (slo.project, slo.service)
                self.service_slo_counts[key] = self.service_slo_counts.get(key, 0) + 1
    
    def _extract_slo_queries(self, spec: Dict[str, Any]) -> Tuple[str, str, str]:
        """Extract queries from SLO spec.
        
        Args:
            spec: SLO specification dictionary
            
        Returns:
            Tuple of (query, numerator_query, denominator_query) as strings
        """
        def _serialize_query(query_obj: Any) -> str:
            """Serialize query object to JSON string if it's a dict, otherwise return as string."""
            if isinstance(query_obj, dict):
                return json.dumps(query_obj, separators=(',', ':'))
            elif query_obj:
                return str(query_obj)
            return ""
        
        def _is_query_dict(obj: Any) -> bool:
            """Check if object is a query dictionary."""
            if not isinstance(obj, dict):
                return False
            query_keys = ["prometheus", "cloudWatch", "splunk", "datadog", 
                         "elasticsearch", "metricSource"]
            return any(key in obj for key in query_keys)
        
        def _extract_query_from_metric(metric: Dict[str, Any]) -> str:
            """Extract query from metric dictionary.
            
            Args:
                metric: Metric dictionary
                
            Returns:
                Serialized query string
            """
            if "query" in metric:
                return _serialize_query(metric["query"])
            elif _is_query_dict(metric):
                return _serialize_query(metric)
            return ""
        
        query = ""
        numerator_query = ""
        denominator_query = ""
        
        objectives = spec.get("objectives", [])
        if not objectives:
            indicator = spec.get("indicator", {})
            if "rawMetric" in indicator:
                raw_metric = indicator.get("rawMetric", {})
                query_obj = raw_metric.get("query", "")
                query = _serialize_query(query_obj)
            elif "thresholdMetric" in indicator:
                threshold_metric = indicator.get("thresholdMetric", {})
                query_obj = threshold_metric.get("query", "")
                query = _serialize_query(query_obj)
            return (query, numerator_query, denominator_query)
        
        for objective in objectives:
            if "composite" in objective:
                continue
                
            if "rawMetric" in objective:
                raw_metric = objective.get("rawMetric", {})
                query_obj = raw_metric.get("query", "")
                query = _serialize_query(query_obj)
            
            elif "thresholdMetric" in objective:
                threshold_metric = objective.get("thresholdMetric", {})
                query_obj = threshold_metric.get("query", "")
                query = _serialize_query(query_obj)
            
            elif "ratioMetric" in objective:
                ratio_metric = objective.get("ratioMetric", {})
                if "good" in ratio_metric:
                    good_metric = ratio_metric.get("good", {})
                    numerator_query = _extract_query_from_metric(good_metric)
                if "total" in ratio_metric:
                    total_metric = ratio_metric.get("total", {})
                    denominator_query = _extract_query_from_metric(total_metric)
            
            elif "countMetrics" in objective:
                count_metrics = objective.get("countMetrics", {})
                if "good" in count_metrics:
                    good_metric = count_metrics.get("good", {})
                    numerator_query = _serialize_query(good_metric)
                if "total" in count_metrics:
                    total_metric = count_metrics.get("total", {})
                    denominator_query = _serialize_query(total_metric)
        
        return (query, numerator_query, denominator_query)
    
    def collect_services(self):
        """Collect all services using sloctl and extract responsibleUsers."""
        print_header("COLLECTING SERVICES")
        print("Collecting services...")
        
        data = self._run_sloctl_command(["get", "services", "-A"])
        self.services = data if data and isinstance(data, list) else []
        
        # Extract responsibleUsers from each service
        for service in self.services:
            metadata = service.get("metadata", {})
            spec = service.get("spec", {})
            project = metadata.get("project", "")
            service_name = metadata.get("name", "")
            
            # Get responsibleUsers from spec
            responsible_users = spec.get("responsibleUsers", [])
            if responsible_users and project and service_name:
                # Handle both list of strings (user IDs) and list of dicts
                # Also handle case where it might be a single value instead of a list
                if not isinstance(responsible_users, list):
                    responsible_users = [responsible_users]
                
                user_ids = []
                for user in responsible_users:
                    if isinstance(user, dict):
                        # Extract user ID from dictionary - prioritize 'id' since that's the common format
                        user_id = user.get("id") or user.get("userId") or user.get("user") or ""
                        if user_id:
                            user_ids.append(str(user_id))
                    elif isinstance(user, str):
                        # Direct user ID string
                        if user.strip():  # Only add non-empty strings
                            user_ids.append(user.strip())
                
                if user_ids:
                    self.service_responsible_users[(project, service_name)] = user_ids
        
        print_colored(f"✓ Collected {len(self.services)} services", colorama.Fore.GREEN)
        print_colored(f"✓ Found {len(self.service_responsible_users)} services with responsible users", colorama.Fore.GREEN)
    
    def fetch_user_information(self):
        """Fetch all users and create lookup dictionary."""
        print_header("FETCHING USER INFORMATION")
        print("Fetching users from Nobl9 API...")
        
        # Determine custom base URL for fetch_users
        # fetch_users expects the base URL without /api (it adds /api/usrmgmt/v2/users internally)
        custom_base_url = None
        if self.is_custom_instance and self.base_url:
            # Remove /api if present, fetch_users will add it
            if self.base_url.endswith('/api'):
                custom_base_url = self.base_url[:-4]  # Remove '/api'
            elif self.base_url.endswith('/'):
                custom_base_url = self.base_url.rstrip('/')
            else:
                custom_base_url = self.base_url
        
        # Use fetch_users from users_basic_v1.2
        users = fetch_users(
            self.access_token,
            self.organization_id,
            self.is_custom_instance,
            custom_base_url
        )
        
        # Create lookup dictionary: user_id -> {name, email}
        for user in users:
            user_id = user.get("userId", "")
            if user_id:
                first_name = user.get("firstName", "")
                last_name = user.get("lastName", "")
                name = f"{first_name} {last_name}".strip() if first_name or last_name else ""
                email = user.get("email", "")
                
                self.user_lookup[user_id] = {
                    "name": name or "Unknown",
                    "email": email or ""
                }
        
        print_colored(f"✓ Created lookup for {len(self.user_lookup)} users", colorama.Fore.GREEN)
    
    def resolve_user_info(self, user_id: str) -> Tuple[str, str]:
        """Resolve user ID to name and email."""
        if not user_id:
            return ("", "")
        
        user_info = self.user_lookup.get(user_id, {})
        return (
            user_info.get("name", user_id),  # Fallback to user_id if not found
            user_info.get("email", "")
        )
    
    def get_service_responsible_users(self, project: str, service_name: str) -> List[Tuple[str, str]]:
        """Get responsible users for a service, resolved to names and emails."""
        if not project or not service_name:
            return []
        
        user_ids = self.service_responsible_users.get((project, service_name), [])
        return [self.resolve_user_info(user_id) for user_id in user_ids]
    
    def _get_base_url_for_links(self) -> str:
        """Get the base URL for hyperlinks, handling custom instances.
        
        Returns:
            Base URL string (e.g., 'https://app.nobl9.com' or 'https://us1.nobl9.com')
        """
        if self.is_custom_instance and self.base_url:
            # Remove /api if present, we want the base URL for web links
            if self.base_url.endswith('/api'):
                return self.base_url[:-4]
            elif self.base_url.endswith('/'):
                return self.base_url.rstrip('/')
            return self.base_url
        return "https://app.nobl9.com"
    
    def _adjust_column_widths(self, worksheet: Any, df: pd.DataFrame) -> None:
        """Helper method to auto-adjust column widths for a worksheet.
        
        Args:
            worksheet: Openpyxl worksheet object
            df: DataFrame containing the data
        """
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(max_length + 2, 50)
    
    def export_to_excel(self):
        """Export SLOs with user information to Excel."""
        print_header("EXPORTING TO EXCEL")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"slo_analysis_{self.context_name}_{timestamp}.xlsx"
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # SLOs with user information (include both regular and composite SLOs)
                all_slos = self.slos + self.composite_slos
                slos_data = []
                for slo in all_slos:
                    user_name, user_email = self.resolve_user_info(slo.created_by)
                    
                    # Get responsible users for this SLO's service
                    responsible_users = self.get_service_responsible_users(slo.project, slo.service)
                    responsible_user_names = [name for name, _ in responsible_users if name]
                    responsible_user_emails = [email for _, email in responsible_users if email]
                    
                    responsible_users_str = ", ".join(responsible_user_names) if responsible_user_names else ""
                    responsible_emails_str = ", ".join(responsible_user_emails) if responsible_user_emails else ""
                    
                    # Determine query columns
                    if slo.numerator_query or slo.denominator_query:
                        query_col = ""
                        numerator_col = slo.numerator_query or ""
                        denominator_col = slo.denominator_query or ""
                    else:
                        query_col = slo.query or ""
                        numerator_col = ""
                        denominator_col = ""
                    
                    alert_policies_str = ", ".join(slo.alert_policies) if slo.alert_policies else "None"
                    
                    slos_data.append({
                        'Name': slo.name,
                        'Display Name': slo.display_name or slo.name,
                        'Type': getattr(slo, 'slo_type', 'Regular'),
                        'Project': slo.project,
                        'Service': slo.service,
                        'Description': slo.description,
                        'Target': slo.target,
                        'Query': query_col,
                        'Numerator Query': numerator_col,
                        'Denominator Query': denominator_col,
                        'Time Window': slo.time_window,
                        'Alert Policies': alert_policies_str,
                        'Health Status': slo.health_status,
                        'Updated At': slo.updated_at,
                        'Created At': slo.created_at,
                        'Created By ID': slo.created_by or "",
                        'Created By Name': user_name,
                        'Created By Email': user_email,
                        'Responsible Users': responsible_users_str,
                        'Responsible User Emails': responsible_emails_str
                    })
                
                if not slos_data:
                    print_colored("Warning: No SLO data to export", colorama.Fore.YELLOW)
                    # Create empty DataFrame with expected columns to avoid "no visible sheet" error
                    slos_df = pd.DataFrame(columns=[
                        'Name', 'Display Name', 'Type', 'Project', 'Service', 'Description', 'Target',
                        'Query', 'Numerator Query', 'Denominator Query', 'Time Window',
                        'Alert Policies', 'Health Status', 'Updated At', 'Created At',
                        'Created By ID', 'Created By Name', 'Created By Email',
                        'Responsible Users', 'Responsible User Emails'
                    ])
                else:
                    slos_df = pd.DataFrame(slos_data)
                    # Replace empty values but don't drop all columns
                    slos_df = slos_df.replace(['', 'None', '[]', 'unknown'], pd.NA)
                    # Only drop columns that are completely empty (all NA)
                    slos_df = slos_df.dropna(axis=1, how='all')
                
                # Ensure we have at least one column
                if slos_df.empty or len(slos_df.columns) == 0:
                    print_colored("Error: No data columns available for export", colorama.Fore.RED)
                    return None
                
                slos_df.to_excel(writer, sheet_name='SLOs', index=False)
                
                worksheet = writer.sheets['SLOs']
                
                # Auto-adjust column widths
                self._adjust_column_widths(worksheet, slos_df)
                
                # Hide specific columns
                columns_to_hide = {
                    'Created By ID',
                    'Target',
                    'Query',
                    'Numerator Query',
                    'Denominator Query',
                    'Time Window',
                    'Alert Policies'
                }
                for idx, col in enumerate(slos_df.columns, 1):
                    if col in columns_to_hide:
                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].hidden = True
                
                # Add hyperlinks to Display Name column
                if self.organization_id:
                    try:
                        base_url = self._get_base_url_for_links()
                        display_name_col_idx = list(slos_df.columns).index('Display Name') + 1
                        for row_idx, slo in enumerate(self.slos, start=2):
                            cell = worksheet.cell(row=row_idx, column=display_name_col_idx)
                            slo_url = f"{base_url}/slo/overview/{slo.project}/{slo.name}?org={self.organization_id}&opt=currentTimeWindow"
                            cell.hyperlink = slo_url
                            cell.style = "Hyperlink"
                    except ValueError:
                        # Display Name column not found, skip hyperlinks
                        pass
                
                # Services sheet
                if self.services:
                    services_data = []
                    for service in self.services:
                        metadata = service.get("metadata", {})
                        spec = service.get("spec", {})
                        project = metadata.get("project", "")
                        service_name = metadata.get("name", "")
                        
                        # Get responsible users for this service
                        responsible_users = self.get_service_responsible_users(project, service_name)
                        responsible_user_names = [name for name, _ in responsible_users if name]
                        responsible_user_emails = [email for _, email in responsible_users if email]
                        
                        # Get SLO count for this service
                        slo_count = self.service_slo_counts.get((project, service_name), 0)
                        
                        services_data.append({
                            'Name': service_name,
                            'SLO Count': slo_count,
                            'Project': project,
                            'Description': spec.get("description", ""),
                            'Responsible Users': ", ".join(responsible_user_names) if responsible_user_names else "",
                            'Responsible User Emails': ", ".join(responsible_user_emails) if responsible_user_emails else ""
                        })
                    
                    if services_data:
                        services_df = pd.DataFrame(services_data)
                        services_df = services_df.replace(['', 'None', '[]'], pd.NA).dropna(axis=1, how='all')
                        if not services_df.empty and len(services_df.columns) > 0:
                            services_df.to_excel(writer, sheet_name='Services', index=False)
                            
                            # Auto-adjust column widths for Services sheet
                            services_worksheet = writer.sheets['Services']
                            self._adjust_column_widths(services_worksheet, services_df)
                            
                            # Center the SLO Count column
                            try:
                                slo_count_col_idx = list(services_df.columns).index('SLO Count') + 1
                                for row in range(1, len(services_df) + 2):  # +2 for header and 1-based indexing
                                    cell = services_worksheet.cell(row=row, column=slo_count_col_idx)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                            except ValueError:
                                # SLO Count column not found, skip centering
                                pass
                            
                            # Add hyperlinks to Name column
                            if self.organization_id:
                                try:
                                    base_url = self._get_base_url_for_links()
                                    name_col_idx = list(services_df.columns).index('Name') + 1
                                    for row_idx, service in enumerate(self.services, start=2):
                                        metadata = service.get("metadata", {})
                                        project = metadata.get("project", "")
                                        service_name = metadata.get("name", "")
                                        if project and service_name:
                                            cell = services_worksheet.cell(row=row_idx, column=name_col_idx)
                                            service_url = f"{base_url}/services/{project}/{service_name}/slos?org={self.organization_id}"
                                            cell.hyperlink = service_url
                                            cell.style = "Hyperlink"
                                except ValueError:
                                    # Name column not found, skip hyperlinks
                                    pass
            
            print_colored(f"✓ Excel file exported: {filename}", colorama.Fore.GREEN)
            return filename
            
        except Exception as e:
            print_colored(f"Error exporting to Excel: {e}", colorama.Fore.RED)
            import traceback
            traceback.print_exc()
            return None


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="SLO Analysis with User Information",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--context", "-c",
        type=str,
        help="Nobl9 context name to use"
    )
    
    args = parser.parse_args()
    
    # Create analyzer
    analyzer = SLOAnalyzer(context_name=args.context)
    
    try:
        # Setup authentication
        analyzer.setup_authentication()
        
        # Collect data
        analyzer.collect_slos()
        analyzer.collect_services()
        
        # Fetch user information
        analyzer.fetch_user_information()
        
        # Export to Excel
        filename = analyzer.export_to_excel()
        
        if filename:
            print_colored(f"\n✓ Analysis complete! File: {filename}", colorama.Fore.GREEN)
        else:
            print_colored("\n✗ Export failed", colorama.Fore.RED)
            sys.exit(1)
            
    except KeyboardInterrupt:
        print_colored("\n\nOperation cancelled by user", colorama.Fore.YELLOW)
        sys.exit(1)
    except Exception as e:
        print_colored(f"\nError: {e}", colorama.Fore.RED)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

