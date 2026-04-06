#!/usr/bin/env python3
"""
Error Budget Burn Report

Analyzes SLO error budget consumption over a specified time period.
Identifies SLOs that have burned through a specified percentage of their error budget.

Features:
- Time period selection (24 hours, 7 days, 14 days, 28 days)
- Customizable error budget drop threshold
- Detailed console report with burn rates and status
- Excel export with hyperlinks to Nobl9 UI
- CSV export option
"""

import json
import sys
import subprocess
import argparse
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from dataclasses import dataclass

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Alignment
    import colorama
    import requests
    colorama.init()
except ImportError as e:
    print(f"Missing required dependencies: {e}")
    print("Install with: pip install pandas openpyxl colorama requests")
    sys.exit(1)

# Import from account_analyzer (same directory)
try:
    from account_analyzer import (
        enhanced_choose_context,
        get_credentials_for_context,
        authenticate
    )
except ImportError as e:
    print(f"Error importing from account_analyzer: {e}")
    sys.exit(1)


def print_colored(text: str, color: str = "") -> None:
    """Print colored text to console."""
    print(f"{color}{text}{colorama.Fore.RESET}")


def print_header(text: str) -> None:
    """Print a formatted header section."""
    print_colored(f"\n{'='*60}", colorama.Fore.CYAN)
    print_colored(text, colorama.Fore.CYAN)
    print_colored('='*60, colorama.Fore.CYAN)


@dataclass
class SLOBudgetInfo:
    """Information about an SLO's error budget status."""
    name: str
    display_name: str
    project: str
    service: str
    slo_type: str
    target: float
    budget_remaining_pct: float
    budget_burned_pct: float
    burn_rate: float
    status: str
    time_window_days: int
    alert_policies: List[str]
    created_at: str


class ErrorBudgetAnalyzer:
    """Analyzes error budget consumption across SLOs."""
    
    def __init__(self, context_name: Optional[str] = None):
        """Initialize the analyzer."""
        self.context_name = context_name
        self.access_token = None
        self.organization_id = None
        self.base_url = "https://app.nobl9.com"
        self.is_custom_instance = False
        self.slos_with_burn: List[SLOBudgetInfo] = []
        self.time_period_hours = 24
        self.budget_drop_threshold = 10.0
        self.total_slos_checked = 0
    
    def setup_authentication(self):
        """Set up authentication and context."""
        print_header("AUTHENTICATION")
        
        if self.context_name:
            credentials = get_credentials_for_context(self.context_name)
            if not credentials:
                print(f"ERROR: Context '{self.context_name}' not found.")
                sys.exit(1)
        else:
            self.context_name, credentials = enhanced_choose_context()
        
        print(f"Selected context: {self.context_name}")
        
        # Authenticate and get token
        self.access_token, self.organization_id, self.is_custom_instance, base_url_from_auth = authenticate(credentials)
        
        if base_url_from_auth:
            self.base_url = base_url_from_auth
        
        print_colored(f"✓ Authenticated successfully", colorama.Fore.GREEN)
        print_colored(f"  Organization: {self.organization_id}", colorama.Fore.GREEN)
        
        # Switch sloctl context
        try:
            subprocess.run(
                ["sloctl", "config", "use-context", self.context_name],
                check=True,
                capture_output=True,
                text=True
            )
            print_colored(f"✓ Switched to context: {self.context_name}", colorama.Fore.GREEN)
        except subprocess.CalledProcessError as e:
            print_colored(f"Warning: Could not switch sloctl context: {e}", colorama.Fore.YELLOW)
    
    def select_time_period(self):
        """Prompt user to select time period for analysis."""
        print_header("TIME PERIOD SELECTION")
        print("Select the time period to analyze error budget burn:")
        print("  [1] 24 hours")
        print("  [2] 7 days")
        print("  [3] 14 days")
        print("  [4] 28 days")
        
        try:
            choice = input("\nSelect time period: ").strip()
            
            time_map = {
                "1": 24,
                "2": 168,  # 7 * 24
                "3": 336,  # 14 * 24
                "4": 672   # 28 * 24
            }
            
            if choice not in time_map:
                print("Invalid selection. Defaulting to 7 days.")
                self.time_period_hours = 168
            else:
                self.time_period_hours = time_map[choice]
            
            days = self.time_period_hours / 24
            print_colored(f"\n✓ Selected: {days:.0f} days ({self.time_period_hours} hours)", colorama.Fore.GREEN)
            
        except KeyboardInterrupt:
            print("\nExiting...")
            sys.exit(0)
    
    def select_budget_threshold(self):
        """Prompt user to input error budget drop threshold."""
        print_header("BUDGET DROP THRESHOLD")
        print("Enter the minimum error budget drop percentage to report.")
        print("Example: Enter '10' to find SLOs that burned 10% or more of their budget")
        
        try:
            threshold_input = input("\nEnter threshold percentage (default: 10): ").strip()
            
            if not threshold_input:
                self.budget_drop_threshold = 10.0
            else:
                try:
                    self.budget_drop_threshold = float(threshold_input)
                    if self.budget_drop_threshold < 0 or self.budget_drop_threshold > 100:
                        print("Invalid threshold. Must be between 0 and 100. Using default: 10%")
                        self.budget_drop_threshold = 10.0
                except ValueError:
                    print("Invalid input. Using default: 10%")
                    self.budget_drop_threshold = 10.0
            
            print_colored(f"\n✓ Threshold set to: {self.budget_drop_threshold}%", colorama.Fore.GREEN)
            
        except KeyboardInterrupt:
            print("\nExiting...")
            sys.exit(0)
    
    def _run_sloctl_command(self, args: List[str]) -> List[Dict]:
        """Run sloctl command and return JSON output."""
        try:
            result = subprocess.run(
                ["sloctl"] + args + ["-o", "json"],
                check=True,
                capture_output=True,
                text=True
            )
            return json.loads(result.stdout)
        except subprocess.CalledProcessError as e:
            print_colored(f"Error running sloctl command: {e}", colorama.Fore.RED)
            return []
        except json.JSONDecodeError as e:
            print_colored(f"Error parsing JSON output: {e}", colorama.Fore.RED)
            return []
    
    def _get_all_slos(self) -> Dict[Tuple[str, str], Dict]:
        """Get all SLO definitions using batch API (v2/slos endpoint).
        
        Returns:
            Dictionary mapping (project, slo_name) to SLO definition
        """
        if self.is_custom_instance and self.base_url:
            api_url = f"{self.base_url}/v2/slos"
        else:
            api_url = f"https://app.nobl9.com/api/v2/slos"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Organization": self.organization_id
        }
        
        status_map = {}
        cursor = None
        page_count = 0
        
        print("Fetching SLO status data from Nobl9 API...")
        
        while True:
            page_count += 1
            params = {"limit": 500}
            if cursor:
                params["cursor"] = cursor
            
            try:
                response = requests.get(api_url, headers=headers, params=params, timeout=30)
                if response.status_code != 200:
                    print_colored(f"API error {response.status_code}: {response.text[:500]}", colorama.Fore.YELLOW)
                    # Try to show response structure for debugging
                    try:
                        error_data = response.json()
                        print_colored(f"Response keys: {list(error_data.keys())}", colorama.Fore.YELLOW)
                    except:
                        pass
                    break
                
                data = response.json()
                
                # Debug: Show response structure
                if page_count == 1:
                    print_colored(f"  API Response keys: {list(data.keys())}", colorama.Fore.CYAN)
                
                # Try different possible key names
                slos = data.get("slos", data.get("data", data.get("items", [])))
                
                print(f"  Page {page_count}: Retrieved {len(slos)} SLOs")
                
                # Build status map
                for slo_item in slos:
                    # Handle different response structures
                    if isinstance(slo_item.get("project"), dict):
                        project = slo_item.get("project", {}).get("name", "")
                    else:
                        project = slo_item.get("project", "")
                    
                    slo_name = slo_item.get("name", "")
                    
                    if project and slo_name:
                        status_map[(project, slo_name)] = slo_item
                
                # Check for next page - try multiple cursor field names
                cursor = data.get("cursor") or data.get("nextCursor") or data.get("next")
                if not cursor:
                    break
                    
            except Exception as e:
                print_colored(f"Error fetching SLO statuses: {e}", colorama.Fore.YELLOW)
                import traceback
                traceback.print_exc()
                break
        
        print_colored(f"✓ Retrieved {len(status_map)} SLO definitions in {page_count} API calls", colorama.Fore.GREEN)
        return status_map
    
    def _get_slo_status(self, project: str, slo_name: str) -> Optional[Dict]:
        """Get status for a single SLO.
        
        Uses the SLO status endpoint which includes error budget and burn rate.
        """
        if self.is_custom_instance and self.base_url:
            # Try both possible paths
            api_url = f"{self.base_url}/slo/v2/status/project/{project}/slo/{slo_name}"
        else:
            api_url = f"https://app.nobl9.com/api/slo/v2/status/project/{project}/slo/{slo_name}"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Organization": self.organization_id
        }
        
        try:
            response = requests.get(api_url, headers=headers, timeout=10)
            if response.status_code == 200:
                return response.json()
            else:
                return None
        except Exception:
            return None
    
    def collect_slo_budget_data(self):
        """Collect SLO data and error budget information."""
        print_header("COLLECTING SLO DATA")
        
        # Step 1: Get all SLO definitions from batch API
        slo_definitions = self._get_all_slos()
        if not slo_definitions:
            print_colored("No SLO definitions retrieved", colorama.Fore.YELLOW)
            return
        
        self.total_slos_checked = len(slo_definitions)
        print(f"\nFetching error budget status for {self.total_slos_checked} SLOs...")
        print_colored("(This requires individual API calls per SLO)", colorama.Fore.CYAN)
        
        # Process each SLO
        processed = 0
        skipped_healthy = 0
        skipped_low_burn = 0
        api_calls_made = 0
        sample_shown = False
        
        for (project, slo_name), slo_item in slo_definitions.items():
            processed += 1
            if processed % 50 == 0:
                print(f"  Analyzed {processed}/{self.total_slos_checked} SLOs...")
            
            # Extract data from v2 API response structure
            display_name = slo_item.get("displayName", slo_name)
            service_obj = slo_item.get("service", {})
            service = service_obj.get("name", "") if isinstance(service_obj, dict) else str(service_obj)
            
            # Check if composite SLO
            objectives = slo_item.get("objectives", [])
            is_composite = False
            for objective in objectives:
                if "composite" in objective and "components" in objective.get("composite", {}):
                    is_composite = True
                    break
            
            slo_type = "Composite" if is_composite else "Regular"
            
            # Get time window in days
            time_windows = slo_item.get("timeWindows", [])
            time_window_days = 0
            if time_windows:
                tw = time_windows[0]
                if tw.get("unit") == "Day":
                    time_window_days = tw.get("count", 0)
                elif tw.get("unit") == "Hour":
                    time_window_days = tw.get("count", 0) / 24
            
            # Get target
            slo_target = 0.0
            if objectives and not is_composite:
                for objective in objectives:
                    if "target" in objective:
                        slo_target = objective.get("target", 0.0)
                        break
            
            # Get status data from individual status endpoint
            api_calls_made += 1
            status_data = self._get_slo_status(project, slo_name)
            
            if not status_data:
                # Skip SLOs we can't get status for
                continue
            
            # Extract error budget and status info from status response
            error_budget = status_data.get("errorBudget", {})
            budget_remaining = error_budget.get("remaining", 100.0)
            burn_rate = status_data.get("burnRate", 0.0)
            health_status = status_data.get("status", "unknown")
            
            # Debug: Show first SLO with budget data
            if not sample_shown and budget_remaining < 100:
                print_colored(f"\nSample SLO with budget burn:", colorama.Fore.YELLOW)
                print_colored(f"  Name: {slo_name}", colorama.Fore.YELLOW)
                print_colored(f"  Budget remaining: {budget_remaining}%", colorama.Fore.YELLOW)
                print_colored(f"  Burn rate: {burn_rate}", colorama.Fore.YELLOW)
                print_colored(f"  Status: {health_status}", colorama.Fore.YELLOW)
                sample_shown = True
            
            # Quick filter: Skip healthy SLOs if threshold is high
            if self.budget_drop_threshold > 20.0 and health_status == "healthy":
                skipped_healthy += 1
                continue
            
            # Calculate budget burned
            budget_burned = (100.0 - budget_remaining)
            
            # Early exit if doesn't meet threshold
            if budget_burned < self.budget_drop_threshold:
                skipped_low_burn += 1
                continue
            
            # Only add if time window is compatible with analysis period
            if time_window_days > 0 and self.time_period_hours / 24 <= time_window_days:
                alert_policies = slo_item.get("alertPolicies", [])
                if isinstance(alert_policies, list):
                    alert_policy_names = [ap.get("name", "") if isinstance(ap, dict) else str(ap) for ap in alert_policies]
                else:
                    alert_policy_names = []
                
                self.slos_with_burn.append(SLOBudgetInfo(
                    name=slo_name,
                    display_name=display_name,
                    project=project,
                    service=service,
                    slo_type=slo_type,
                    target=slo_target,
                    budget_remaining_pct=budget_remaining,
                    budget_burned_pct=budget_burned,
                    burn_rate=burn_rate,
                    status=health_status,
                    time_window_days=time_window_days,
                    alert_policies=alert_policy_names,
                    created_at=slo_item.get("createdAt", "")
                ))
        
        # Sort by budget burned descending
        self.slos_with_burn.sort(key=lambda x: x.budget_burned_pct, reverse=True)
        
        print_colored(f"\n✓ Analyzed {self.total_slos_checked} SLOs", colorama.Fore.GREEN)
        print_colored(f"  - Made {api_calls_made} status API calls", colorama.Fore.CYAN)
        if skipped_healthy > 0:
            print_colored(f"  - Skipped {skipped_healthy} healthy SLOs (optimization)", colorama.Fore.CYAN)
        if skipped_low_burn > 0:
            print_colored(f"  - Skipped {skipped_low_burn} SLOs below threshold", colorama.Fore.CYAN)
        print_colored(f"✓ Found {len(self.slos_with_burn)} SLOs with {self.budget_drop_threshold}%+ budget burn", colorama.Fore.GREEN)
    
    def display_console_report(self):
        """Display error budget burn report in console."""
        if not self.slos_with_burn:
            print_colored(f"\n✓ No SLOs found with {self.budget_drop_threshold}%+ error budget burn", colorama.Fore.GREEN)
            return
        
        print_header("ERROR BUDGET BURN REPORT")
        days = self.time_period_hours / 24
        print_colored(f"Time Period: {days:.0f} days", colorama.Fore.WHITE)
        print_colored(f"Budget Drop Threshold: {self.budget_drop_threshold}%", colorama.Fore.WHITE)
        print_colored(f"Organization: {self.organization_id}", colorama.Fore.WHITE)
        
        # Summary
        print_header("SUMMARY")
        print_colored(f"Total SLOs Analyzed: {self.total_slos_checked}", colorama.Fore.WHITE)
        print_colored(f"SLOs with High Burn: {len(self.slos_with_burn)}", colorama.Fore.RED)
        
        # Calculate statistics
        avg_burn = sum(slo.budget_burned_pct for slo in self.slos_with_burn) / len(self.slos_with_burn)
        max_burn = max(slo.budget_burned_pct for slo in self.slos_with_burn)
        avg_burn_rate = sum(slo.burn_rate for slo in self.slos_with_burn) / len(self.slos_with_burn)
        
        print_colored(f"Average Budget Burned: {avg_burn:.1f}%", colorama.Fore.WHITE)
        print_colored(f"Maximum Budget Burned: {max_burn:.1f}%", colorama.Fore.WHITE)
        print_colored(f"Average Burn Rate: {avg_burn_rate:.2f}x", colorama.Fore.WHITE)
        
        # Group by project
        print_header("SLOS WITH HIGH ERROR BUDGET BURN")
        
        project_slos = {}
        for slo in self.slos_with_burn:
            if slo.project not in project_slos:
                project_slos[slo.project] = []
            project_slos[slo.project].append(slo)
        
        for project in sorted(project_slos.keys()):
            print_colored(f"\nProject: {project}", colorama.Fore.CYAN)
            print_colored("-" * 60, colorama.Fore.CYAN)
            
            for slo in project_slos[project]:
                # Color code based on burn severity
                if slo.budget_burned_pct >= 50:
                    color = colorama.Fore.RED
                    severity = "CRITICAL"
                elif slo.budget_burned_pct >= 25:
                    color = colorama.Fore.YELLOW
                    severity = "WARNING"
                else:
                    color = colorama.Fore.WHITE
                    severity = "NOTICE"
                
                print_colored(f"\n  [{severity}] {slo.display_name or slo.name}", color)
                print_colored(f"    Service: {slo.service}", colorama.Fore.WHITE)
                print_colored(f"    Type: {slo.slo_type}", colorama.Fore.WHITE)
                print_colored(f"    Budget Remaining: {slo.budget_remaining_pct:.1f}%", colorama.Fore.WHITE)
                print_colored(f"    Budget Burned: {slo.budget_burned_pct:.1f}%", color)
                print_colored(f"    Burn Rate: {slo.burn_rate:.2f}x", color)
                print_colored(f"    Status: {slo.status}", colorama.Fore.WHITE)
                print_colored(f"    Time Window: {slo.time_window_days} days", colorama.Fore.WHITE)
        
        # Top burners
        print_header("TOP 10 ERROR BUDGET BURNERS")
        print()
        print_colored(f"  {'Rank':<6} {'Budget Burned':<15} {'Burn Rate':<12} {'SLO Name'}", colorama.Fore.CYAN)
        print_colored(f"  {'-'*6} {'-'*15} {'-'*12} {'-'*40}", colorama.Fore.CYAN)
        
        for idx, slo in enumerate(self.slos_with_burn[:10], 1):
            if slo.budget_burned_pct >= 50:
                color = colorama.Fore.RED
            elif slo.budget_burned_pct >= 25:
                color = colorama.Fore.YELLOW
            else:
                color = colorama.Fore.WHITE
            
            print_colored(
                f"  {idx:<6} {slo.budget_burned_pct:>6.1f}%{'':<8} {slo.burn_rate:>5.2f}x{'':<6} {slo.display_name or slo.name}",
                color
            )
    
    def _get_base_url_for_links(self) -> str:
        """Get the base URL for Nobl9 UI links."""
        if self.is_custom_instance and self.base_url:
            if "us1.nobl9.com" in self.base_url:
                return "https://us1.nobl9.com"
            elif "cg1.nobl9.com" in self.base_url:
                return "https://cg1.nobl9.com"
        return "https://app.nobl9.com"
    
    def _adjust_column_widths(self, worksheet, df):
        """Auto-adjust column widths in Excel."""
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(max_length + 2, 50)
    
    def export_to_excel(self):
        """Export error budget burn data to Excel."""
        if not self.slos_with_burn:
            print_colored("\nNo data to export", colorama.Fore.YELLOW)
            return None
        
        print_header("EXPORTING TO EXCEL")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        days = int(self.time_period_hours / 24)
        filename = f"error_budget_report_{self.context_name}_{days}days_{timestamp}.xlsx"
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Prepare data
                data = []
                for slo in self.slos_with_burn:
                    # Determine severity
                    if slo.budget_burned_pct >= 50:
                        severity = "CRITICAL"
                    elif slo.budget_burned_pct >= 25:
                        severity = "WARNING"
                    else:
                        severity = "NOTICE"
                    
                    alert_policies_str = ", ".join(slo.alert_policies) if slo.alert_policies else "None"
                    
                    data.append({
                        'SLO Name': slo.name,
                        'Display Name': slo.display_name or slo.name,
                        'Severity': severity,
                        'Project': slo.project,
                        'Service': slo.service,
                        'Type': slo.slo_type,
                        'Budget Remaining %': round(slo.budget_remaining_pct, 2),
                        'Budget Burned %': round(slo.budget_burned_pct, 2),
                        'Burn Rate': round(slo.burn_rate, 2),
                        'Target': slo.target,
                        'Status': slo.status,
                        'Time Window (days)': slo.time_window_days,
                        'Alert Policies': alert_policies_str,
                        'Created At': slo.created_at
                    })
                
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name='Error Budget Report', index=False)
                
                worksheet = writer.sheets['Error Budget Report']
                
                # Auto-adjust column widths
                self._adjust_column_widths(worksheet, df)
                
                # Add hyperlinks to Display Name column
                if self.organization_id:
                    try:
                        base_url = self._get_base_url_for_links()
                        display_name_col_idx = list(df.columns).index('Display Name') + 1
                        for row_idx, slo in enumerate(self.slos_with_burn, start=2):
                            cell = worksheet.cell(row=row_idx, column=display_name_col_idx)
                            slo_url = f"{base_url}/slo/overview/{slo.project}/{slo.name}?org={self.organization_id}&opt=currentTimeWindow"
                            cell.hyperlink = slo_url
                            cell.style = "Hyperlink"
                    except ValueError:
                        pass
                
                # Center numeric columns
                for col_name in ['Budget Remaining %', 'Budget Burned %', 'Burn Rate', 'Time Window (days)']:
                    try:
                        col_idx = list(df.columns).index(col_name) + 1
                        for row_idx in range(2, len(self.slos_with_burn) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(horizontal='center')
                    except ValueError:
                        pass
            
            print_colored(f"✓ Exported to: {filename}", colorama.Fore.GREEN)
            return filename
            
        except Exception as e:
            print_colored(f"Error exporting to Excel: {e}", colorama.Fore.RED)
            return None
    
    def export_to_csv(self):
        """Export error budget burn data to CSV."""
        if not self.slos_with_burn:
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        days = int(self.time_period_hours / 24)
        filename = f"error_budget_report_{self.context_name}_{days}days_{timestamp}.csv"
        
        try:
            data = []
            for slo in self.slos_with_burn:
                if slo.budget_burned_pct >= 50:
                    severity = "CRITICAL"
                elif slo.budget_burned_pct >= 25:
                    severity = "WARNING"
                else:
                    severity = "NOTICE"
                
                alert_policies_str = ", ".join(slo.alert_policies) if slo.alert_policies else "None"
                
                data.append({
                    'SLO Name': slo.name,
                    'Display Name': slo.display_name or slo.name,
                    'Severity': severity,
                    'Project': slo.project,
                    'Service': slo.service,
                    'Type': slo.slo_type,
                    'Budget Remaining %': round(slo.budget_remaining_pct, 2),
                    'Budget Burned %': round(slo.budget_burned_pct, 2),
                    'Burn Rate': round(slo.burn_rate, 2),
                    'Target': slo.target,
                    'Status': slo.status,
                    'Time Window (days)': slo.time_window_days,
                    'Alert Policies': alert_policies_str,
                    'Created At': slo.created_at
                })
            
            df = pd.DataFrame(data)
            df.to_csv(filename, index=False)
            
            print_colored(f"✓ Exported to CSV: {filename}", colorama.Fore.GREEN)
            return filename
            
        except Exception as e:
            print_colored(f"Error exporting to CSV: {e}", colorama.Fore.RED)
            return None


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description="Nobl9 Error Budget Burn Report - Analyze SLO error budget consumption"
    )
    parser.add_argument(
        "--context",
        help="Nobl9 context name (if not provided, will prompt for selection)"
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="Export to CSV instead of Excel"
    )
    
    args = parser.parse_args()
    
    try:
        analyzer = ErrorBudgetAnalyzer(context_name=args.context)
        
        # Setup
        analyzer.setup_authentication()
        analyzer.select_time_period()
        analyzer.select_budget_threshold()
        
        # Collect data
        analyzer.collect_slo_budget_data()
        
        # Display report
        analyzer.display_console_report()
        
        # Export
        if args.csv:
            analyzer.export_to_csv()
        else:
            analyzer.export_to_excel()
        
        print_colored("\n✓ Analysis complete!", colorama.Fore.GREEN)
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print_colored(f"\nError: {e}", colorama.Fore.RED)
        sys.exit(1)


if __name__ == "__main__":
    main()
